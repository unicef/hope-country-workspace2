import io
import os
from typing import TYPE_CHECKING

import openpyxl
import pytest
import xlsxwriter
from django.urls import reverse
from testutils.factories import FlexFieldFactory
from testutils.utils import select_office
from webtest import Checkbox, Upload

from country_workspace.state import state
from country_workspace.workspaces.admin.cleaners.bulk_update import TYPES, create_xls_importer

if TYPE_CHECKING:
    from django_webtest import DjangoTestApp
    from django_webtest.pytest_plugin import MixinWithInstanceVariables
    from pytest_django.fixtures import SettingsWrapper

    from country_workspace.models import AsyncJob
    from country_workspace.workspaces.models import CountryHousehold

pytestmark = [pytest.mark.admin, pytest.mark.smoke, pytest.mark.django_db]


@pytest.fixture
def office():
    from testutils.factories import OfficeFactory

    co = OfficeFactory()
    state.tenant = co
    return co


@pytest.fixture
def program(office, force_migrated_records, household_checker, individual_checker):
    from testutils.factories import CountryProgramFactory

    return CountryProgramFactory(
        country_office=office,
        household_checker=household_checker,
        individual_checker=individual_checker,
        household_columns="__str__\nid\nxx",
        individual_columns="__str__\nid\nxx",
    )


@pytest.fixture
def household(program):
    from testutils.factories import CountryHouseholdFactory

    return CountryHouseholdFactory(
        batch__program=program, batch__country_office=program.country_office, flex_fields={"size": 5}
    )


@pytest.fixture(scope="session")
def celery_config():
    return {"broker_url": os.environ["CELERY_BROKER_URL"], "result_backend": os.environ["CELERY_BROKER_URL"]}


@pytest.fixture(scope="session")
def celery_worker_parameters():
    return {
        "shutdown_timeout": 60,
    }


@pytest.fixture
def celery_app(celery_app):
    return celery_app


@pytest.fixture
def app(django_app_factory: "MixinWithInstanceVariables") -> "DjangoTestApp":
    from testutils.factories import SuperUserFactory

    django_app = django_app_factory(csrf_checks=False)
    admin_user = SuperUserFactory(username="superuser")
    django_app.set_user(admin_user)
    django_app._user = admin_user
    return django_app


@pytest.mark.parametrize(("field", "validator"), list(TYPES.items()))
def test_validator(field, validator):
    flex_field = FlexFieldFactory(definition__field_type=field, definition__attrs={"choices": [("a", "A")]})
    assert validator(flex_field)()


def test_create_xls_importer(household: "CountryHousehold", force_migrated_records):
    selected_fields = [
        "id",
        "gender",
        "given_name",
        "role",
        "relationship",
        "first_registration_date",
        "birth_date",
        "disability",
    ]
    ret, __ = create_xls_importer(
        household.members.all(),
        household.program,
        selected_fields,
    )
    workbook = openpyxl.load_workbook(io.BytesIO(ret.getvalue()))
    sheet = workbook.worksheets[0]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert headers == selected_fields


def test_bulk_update_export(
    app: "DjangoTestApp", force_migrated_records, settings: "SettingsWrapper", household: "CountryHousehold"
) -> None:
    url = reverse("workspace:workspaces_countryindividual_changelist")
    settings.CELERY_TASK_ALWAYS_EAGER = True

    selected_fields = [
        "birth_date",
        "disability",
        "first_registration_date",
        "gender",
        "given_name",
        "relationship",
        "role",
    ]
    with select_office(app, household.country_office, household.program):
        res = app.get(url)
        form = res.forms["changelist-form"]
        form["action"] = "bulk_update_export"
        form.set("_selected_action", True, index=0)
        res = form.submit()

        form = res.forms["bulk-update-form"]
        for i in range(len(form.fields.get("fields"))):
            target: Checkbox = form.fields.get("fields")[i]
            if target._value in selected_fields:
                target.checked = True
        res = form.submit("_export")

        assert res.status_code == 302
        job: AsyncJob = household.program.jobs.first()
        job.queue()


@pytest.fixture
def data(household):
    buff = io.BytesIO()
    workbook = xlsxwriter.Workbook(buff)
    worksheet = workbook.add_worksheet()
    fields = ["gender", "given_name", "role", "relationship", "first_registration_date", "birth_date", "disability"]
    worksheet.write(0, 0, "id")
    for i, c in enumerate(fields, 1):
        worksheet.write(0, i, c)

    for row, e in enumerate(household.members.all(), 1):
        worksheet.write(row, 0, e.id)
        for col, c in enumerate(fields, 1):
            worksheet.write(row, col, f"{c}_{col}")
    workbook.close()
    buff.seek(0)
    return buff, household


def test_bulk_update_import(app: "DjangoTestApp", force_migrated_records, settings: "SettingsWrapper", data) -> None:
    xls, household = data
    url = reverse("workspace:workspaces_countryprogram_change", args=[household.program.pk])
    settings.CELERY_TASK_ALWAYS_EAGER = True
    with select_office(app, household.country_office, household.program):
        res = app.get(url)
        res = res.click("Update Records")
        res.forms["bulk-update-form"]["description"] = "Bulk update from file"
        res.forms["bulk-update-form"]["target"] = "ind"
        res.forms["bulk-update-form"]["file"] = Upload("file.xlsx", xls.read())
        res = res.forms["bulk-update-form"].submit("_import")
        assert res.status_code == 302
        job: AsyncJob = household.program.jobs.first()
        assert job
        assert household.members.filter(flex_fields__given_name="given_name_2").exists()
